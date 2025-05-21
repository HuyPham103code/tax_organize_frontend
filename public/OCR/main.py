# from azure.core.credentials import AzureKeyCredential
# from azure.ai.documentintelligence import DocumentIntelligenceClient
# import ast
# import openpyxl
# from openpyxl.utils import get_column_letter
# import json
# import re

# endpoint_pro = "https://henryphamocr.cognitiveservices.azure.com/"
# key_pro = "BgokkCOFCdGxs3Nn0u03fOLproraZm4rQzAHXpxunThC8PSnzuR7JQQJ99BDACYeBjFXJ3w3AAALACOG4v7Z"
# filepath = r"C:\Users\Admin\Downloads\20181231 ADP Year End W2.pdf"

# print("starting")

# #=======================================extract data======================================
# client = DocumentIntelligenceClient(endpoint_pro, AzureKeyCredential(key_pro))

# with open(filepath, "rb") as f:
#     poller = client.begin_analyze_document("prebuilt-tax.us", body=f)
#     result = poller.result()

# with open("test_w2.txt", "w", encoding="utf-8") as out_file:
#     for i, doc in enumerate(result.documents):
#         out_file.write(f"--- Document {i+1} ---\n")
#         out_file.write(f"Type: {doc.doc_type}\n")
#         out_file.write("All Fields:\n")
#         out_file.write(f"{doc.fields}\n\n")

# print("✅ step1")


# #=======================================transform to json=================================

# txt_path = "test_w2.txt"  # <- Đặt tên file .txt bạn đã lưu ở đây

# documents = []
# with open(txt_path, "r", encoding="utf-8") as file:
#     lines = file.readlines()

# current_doc = {}
# inside_fields = False
# field_lines = []

# for line in lines:
#     line = line.strip()

#     # Bắt đầu document mới
#     if line.startswith("--- Document"):
#         if field_lines:
#             try:
#                 all_fields_str = "\n".join(field_lines)
#                 parsed_dict = ast.literal_eval(all_fields_str)
#                 current_doc["fields"] = parsed_dict
#             except Exception as e:
#                 current_doc["fields"] = {"error": str(e)}
#             documents.append(current_doc)
#             field_lines = []

#         current_doc = {"document_number": int(re.findall(r'\d+', line)[0])}

#     elif line.startswith("Type:"):
#         current_doc["type"] = line.replace("Type:", "").strip()

#     elif line.startswith("All Fields:"):
#         inside_fields = True
#         field_lines = []

#     elif inside_fields:
#         if line.startswith("--- Document") or line.startswith("Type:"):
#             inside_fields = False
#         else:
#             field_lines.append(line)

# # Thêm document cuối cùng nếu còn
# if field_lines:
#     try:
#         all_fields_str = "\n".join(field_lines)
#         current_doc["fields"] = ast.literal_eval(all_fields_str)
#     except Exception as e:
#         current_doc["fields"] = {"error": str(e)}
#     documents.append(current_doc)


# print(f"✅ step2")

# #========================================extract JSON detail==========================

# def extract_fields(doc):
#     fields = doc.get("fields", {})
#     employer = fields.get("Employer", {}).get("valueObject", {})
#     employer_address = employer.get("Address", {}).get("valueAddress", {})

#     state_infos = fields.get("StateTaxInfos", {}).get("valueArray", [])
#     local_infos = fields.get("LocalTaxInfos", {}).get("valueArray", [])

#     def safe_get(d, key, default=""):
#         return d.get(key, {}).get("valueString") or d.get(key, {}).get("valueNumber") or ""

#     result = {
#         "document_number": doc.get("document_number", ""),
#         "Employer_IdNumber": safe_get(employer, "IdNumber"),
#         "Employer_Name": safe_get(employer, "Name"),
#         # Tách từng phần địa chỉ
#         # "Employer_Address_HouseNumber": employer_address.get("houseNumber", ""),
#         # "Employer_Address_Road": employer_address.get("road", ""),
#         "Employer_Address_StreetAddress": employer_address.get("streetAddress", ""),
#         "Employer_Address_City": employer_address.get("city", ""),
#         "Employer_Address_State": employer_address.get("state", ""),
#         "Employer_Address_PostalCode": employer_address.get("postalCode", ""),
#         # "Employer_Address_Unit": employer_address.get("unit", "")
#     }

#     # Các trường số/chuỗi khác
#     keys = [
#         "ControlNumber",
#         "WagesTipsAndOtherCompensation",
#         "FederalIncomeTaxWithheld",
#         "SocialSecurityWages",
#         "SocialSecurityTaxWithheld",
#         "MedicareWagesAndTips",
#         "MedicareTaxWithheld",
#         "SocialSecurityTips",
#         "AllocatedTips",
#         "DependentCareBenefits",
#         "NonQualifiedPlans",
#         "IsStatutoryEmployee",
#         "IsRetirementPlan",
#         "IsThirdPartySickPay"
#     ]
#     for key in keys:
#         result[key] = safe_get(fields, key)

#     # Lấy StateTaxInfos (nếu có)
#     result["StateTaxInfos"] = []
#     for s in state_infos:
#         val = s.get("valueObject", {})
#         result["StateTaxInfos"].append({
#             "State": safe_get(val, "State"),
#             "StateWagesTipsEtc": safe_get(val, "StateWagesTipsEtc"),
#             "StateIncomeTax": safe_get(val, "StateIncomeTax")
#         })

#     # Lấy LocalTaxInfos (nếu có)
#     result["LocalTaxInfos"] = []
#     for l in local_infos:
#         val = l.get("valueObject", {})
#         result["LocalTaxInfos"].append({
#             "LocalWagesTipsEtc": safe_get(val, "LocalWagesTipsEtc"),
#             "LocalIncomeTax": safe_get(val, "LocalIncomeTax")
#         })

#     return result

# # Áp dụng trích xuất cho tất cả document
# filtered_result = [extract_fields(doc) for doc in documents]


# print(f"✅ step3")

# #=====================================generate excel file===============================
# import openpyxl
# from openpyxl.utils import get_column_letter

# # ======= CẤU HÌNH ========
# template_excel_path = "w2_format.xlsm"        # file Excel mẫu có sẵn
# output_excel_path = "test_w2.xlsx"   # file output để lưu

# sheet_name = "W-2 Summary"                    # worksheet cần fill

# # ======= LOAD JSON ========
# data = filtered_result

# # ======= MỞ EXCEL TEMPLATE ========
# wb = openpyxl.load_workbook(template_excel_path)
# ws = wb[sheet_name]

# # ======= CÁC FIELD CẦN FILL ========
# main_fields = [
#     "Employer_IdNumber", "Employer_Name", "Employer_Address_StreetAddress", "Employer_Address_City",
#     "Employer_Address_State", "Employer_Address_PostalCode", "ControlNumber",
#     "WagesTipsAndOtherCompensation", "FederalIncomeTaxWithheld", "SocialSecurityWages",
#     "SocialSecurityTaxWithheld", "MedicareWagesAndTips", "MedicareTaxWithheld",
#     "SocialSecurityTips", "AllocatedTips", "DependentCareBenefits", "NonQualifiedPlans"
# ]
# main_start_row = 9

# boolean_fields = ["IsStatutoryEmployee", "IsRetirementPlan", "IsThirdPartySickPay"]
# boolean_start_row = 36

# state_tax_fields = ["State", "StateWagesTipsEtc", "StateIncomeTax"]
# local_tax_fields = ["LocalWagesTipsEtc", "LocalIncomeTax"]
# tax_start_row = 40

# # ======= GHI DỮ LIỆU ========
# for i, doc in enumerate(data):
#     col_letter = get_column_letter(i + 2)  # B = 2, C = 3, ...

#     # Ghi các field chính
#     for j, field in enumerate(main_fields):
#         ws[f"{col_letter}{main_start_row + j}"] = doc.get(field, "")

#     # Ghi boolean flags
#     for j, field in enumerate(boolean_fields):
#         ws[f"{col_letter}{boolean_start_row + j}"] = doc.get(field, "")

#     # Ghi state tax (chỉ object đầu tiên nếu có)
#     state_infos = doc.get("StateTaxInfos")
#     state_info = state_infos[0] if isinstance(state_infos, list) and state_infos else {}
#     for j, field in enumerate(state_tax_fields):
#         ws[f"{col_letter}{tax_start_row + j}"] = state_info.get(field, "")

#     # Ghi local tax (chỉ object đầu tiên nếu có)
#     local_infos = doc.get("LocalTaxInfos")
#     local_info = local_infos[0] if isinstance(local_infos, list) and local_infos else {}
#     for j, field in enumerate(local_tax_fields):
#         ws[f"{col_letter}{tax_start_row + len(state_tax_fields) + j}"] = local_info.get(field, "")

# # ======= LƯU FILE MỚI ========
# wb.save(output_excel_path)
# print(f"✅ step4")




from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
import json
import re
import ast
import json
import openpyxl


endpoint_pro = "https://henryphamocr.cognitiveservices.azure.com/"
key_pro = "BgokkCOFCdGxs3Nn0u03fOLproraZm4rQzAHXpxunThC8PSnzuR7JQQJ99BDACYeBjFXJ3w3AAALACOG4v7Z"
filepath = r"C:\Users\Admin\Downloads\1099-div.png"

document_intelligence_client  = DocumentIntelligenceClient(
    endpoint=endpoint_pro, credential=AzureKeyCredential(key_pro)
)

with open(filepath, "rb") as f:
    poller = document_intelligence_client.begin_analyze_document(
        "prebuilt-tax.us.1099DIV", body=f
    )
tax1040 = poller.result()

with open("1099_DIV_result.txt", "w", encoding="utf-8") as out_file:
    for i, doc in enumerate(tax1040.documents):
        out_file.write(f"--- Document {i+1} ---\n")
        out_file.write(f"Type: {doc.doc_type}\n")
        out_file.write("All Fields:\n")
        out_file.write(f"{doc.fields}\n\n")

print("✅ step1: 1099_DIV_result.txt")



txt_path = "1099_DIV_result.txt"  # <- Đặt tên file .txt bạn đã lưu ở đây

documents = []
with open(txt_path, "r", encoding="utf-8") as file:
    lines = file.readlines()

current_doc = {}
inside_fields = False
field_lines = []

for line in lines:
    line = line.strip()

    # Bắt đầu document mới
    if line.startswith("--- Document"):
        if field_lines:
            try:
                all_fields_str = "\n".join(field_lines)
                parsed_dict = ast.literal_eval(all_fields_str)
                current_doc["fields"] = parsed_dict
            except Exception as e:
                current_doc["fields"] = {"error": str(e)}
            documents.append(current_doc)
            field_lines = []

        current_doc = {"document_number": int(re.findall(r'\d+', line)[0])}

    elif line.startswith("Type:"):
        current_doc["type"] = line.replace("Type:", "").strip()

    elif line.startswith("All Fields:"):
        inside_fields = True
        field_lines = []

    elif inside_fields:
        if line.startswith("--- Document") or line.startswith("Type:"):
            inside_fields = False
        else:
            field_lines.append(line)

# Thêm document cuối cùng nếu còn
if field_lines:
    try:
        all_fields_str = "\n".join(field_lines)
        current_doc["fields"] = ast.literal_eval(all_fields_str)
    except Exception as e:
        current_doc["fields"] = {"error": str(e)}
    documents.append(current_doc)

print(f"✅ step2: Done")

# Hàm helper an toàn
def safe_get(d, key, default=""):
    return d.get(key, {}).get("valueString") or d.get(key, {}).get("valueNumber") or d.get(key, {}).get("valueBoolean") or default

def extract_fields(doc):
    fields = doc.get("fields", {})
    document_number = doc.get("document_number", "")

    payer = fields.get("Payer", {}).get("valueObject", {})
    recipient = fields.get("Recipient", {}).get("valueObject", {})
    payer_address = payer.get("Address", {}).get("valueAddress", {})
    transactions = fields.get("Transactions", {}).get("valueArray", [])
    first_txn = transactions[0].get("valueObject", {}) if transactions else {}

    state_taxes = first_txn.get("StateTaxesWithheld", {}).get("valueArray", [])
    first_state_tax = state_taxes[0].get("valueObject", {}) if state_taxes else {}

    return {
        "document_number": document_number,
        "Payer_Name": safe_get(payer, "Name"),
        "Payer_TIN": safe_get(payer, "TIN"),
        "Account_Number": safe_get(recipient, "AccountNumber"),

        "Payer_PostalCode": payer_address.get("postalCode", ""),
        "Payer_City": payer_address.get("city", ""),
        "Payer_State": payer_address.get("state", ""),
        
        "Box1a": safe_get(first_txn, "Box1a"),
        "Box1b": safe_get(first_txn, "Box1b"),
        "Box2a": safe_get(first_txn, "Box2a"),
        "Box2b": safe_get(first_txn, "Box2b"),
        "Box2c": safe_get(first_txn, "Box2c"),
        "Box2d": safe_get(first_txn, "Box2d"),

        "Box7": safe_get(first_txn, "Box7"),
        
        "Box9": safe_get(first_txn, "Box9"),
        "Box10": safe_get(first_txn, "Box10"),
        "Box12": safe_get(first_txn, "Box12"),
        "Box13": safe_get(first_txn, "Box13")
    }

# Thực hiện lọc
filtered_result = [extract_fields(doc) for doc in documents]

print(f"✅ Step3: Done")

# ======= CẤU HÌNH ========
template_excel_path = "1099_div_format.xlsx"  # file Excel mẫu
output_excel_path = "1099_div_filled2.xlsx"  # file Excel sau khi fill

sheet_name = "Brokerage Summary (Sch B, D)"

# ======= LOAD DỮ LIỆU ========
data = filtered_result

# ======= MỞ EXCEL TEMPLATE ========
wb = openpyxl.load_workbook(template_excel_path)
ws = wb[sheet_name]

# ======= GHI DỮ LIỆU CHO MỖI DOCUMENT ========
for i, doc in enumerate(data):
    col_letter = openpyxl.utils.get_column_letter(i + 4)  # D = 4, E = 5,...

    # Ghi thông tin chung
    ws[f"{col_letter}8"] = "2-DIV"
    ws[f"{col_letter}9"] = doc.get("Payer_Name", "")
    ws[f"{col_letter}10"] = doc.get("Payer_TIN", "")
    ws[f"{col_letter}11"] = doc.get("Account_Number", "")

    # Box1a → Box2d (ô 35-40)
    ws[f"{col_letter}35"] = doc.get("Box1a", "")
    ws[f"{col_letter}36"] = doc.get("Box1b", "")
    ws[f"{col_letter}37"] = doc.get("Box2a", "")
    ws[f"{col_letter}38"] = doc.get("Box2b", "")
    ws[f"{col_letter}39"] = doc.get("Box2c", "")
    ws[f"{col_letter}40"] = doc.get("Box2d", "")

    # Box7
    ws[f"{col_letter}42"] = doc.get("Box7", "")

    # Box9 → Box13 (ô 45-48)
    ws[f"{col_letter}45"] = doc.get("Box9", "")
    ws[f"{col_letter}46"] = doc.get("Box10", "")
    ws[f"{col_letter}47"] = doc.get("Box12", "")
    ws[f"{col_letter}48"] = doc.get("Box13", "")

# ======= LƯU FILE ========
wb.save(output_excel_path)
print(f"✅ Đã fill dữ liệu vào file: {output_excel_path}")

